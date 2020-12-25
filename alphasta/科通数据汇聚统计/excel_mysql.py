# -*- coding: utf-8 -*-

import datetime
import openpyxl
import pymysql
import time


excel_path= r'D:\guohui\汇聚2020\汇聚2020统计.xlsx'
# excel_path = r'C:\Users\bench\Desktop\2020科通数据汇聚\汇聚2020统计.xlsx'
# mysql_host = '192.168.23.112'
# mysql_host = '13.32.4.170'
# mysql_host = '192.168.0.122'
mysql_host = '20.32.0.49'
# mysql_host = '192.168.0.116'

nonExam_list = ['反恐']
depart_dict = {'治安':1,'技侦':2,'经侦':3,'网安':4,'交管':5,'站前分局':6,'桥西':7,'新华':8,'裕华':9,'长安':10,'藁城':11,'栾城':12,'鹿泉':13,'化工园':14,'矿区':15,'高新':16,'深泽':17,'正定':18,'元氏':19,'高邑':20,'赞皇':21,'赵县':22,'晋州':23,'井陉':24,'新乐':25,'无极':26,'平山':27,'灵寿':28,'行唐':29,'轨道交通分局':57, '反恐':22223}
tables_dict = {'保险理赔信息': '33', '保险信息': '33', '毕业生信息': '38', '残疾人登记信息': '42', '产品纠纷相关人员信息': '99', '慈善捐款信息': '42', '出租车登记信息': '55', '出租车轨迹信息': '55', '出租房屋登记信息': '53', '财政一卡通': '99', '低保家庭信息': '99', '低保人员信息': '99', '典当信息': '37', '电力用户登记信息': '48', '单位从业人员信息': '41', '导游信息': '36', '电子邮箱登记信息': '53', '法律援助对象信息': '54', '房屋登记信息': '30', '房屋交易信息': '30', '抚恤信息': '99', '工地流动人员信息': '41', '公共自行车卡信息': '99', '公交从业人员信息': '55', '公交一卡通充值信息': '55', '公交一卡通交易信息': '55', '公交一卡通退卡信息': '55', '公交一卡通用户信息': '55', '高考核查人员信息': '99', '公路客运购票信息': '56', '供暖信息': '50', '高速公路车辆计重收费通行信息': '56', '高速公路车辆通行信息': '99', '工伤人员信息': '99', '公务人员信息': '41', '购物网站登记信息': '53', '公租房入住信息': '30', '高自考信息': '38', '话单流水信息': '52', '火化信息': '42', '会员卡登记信息': '34', '会员卡刷卡信息': '34', '婚姻状况信息': '42', '军队转业干部': '99', '军队自主择业人员信息': '99', '结婚登记信息': '42', '计划生育服务对象登记信息': '40', '驾驶证约考信息': '99', '交通运输违法案件信息': '35', '机修业车辆交接信息': '99', '机修业/二手车车辆信息': '37', '机修业/二手车从业人员信息': '37', '交友征婚登记信息': '53', '家政服务登记信息': '53', '救助信息': '42', '快递信息': '35', '宽带上网用户登记信息': '53', '客运车辆登记信息': '56', '离婚登记信息': '42', '论坛登记信息': '53', '劳动用工市场登记信息': '41', '旅行社接待旅客信息': '36', '门诊登记信息': '39', '汽车租赁信息': '37', '全国旅馆VIP会员': '99', 'QQ登记信息': '53', 'QQ好友信息': '53', 'QQ群信息': '53', '求租房屋登记信息': '53', '人才市场登记信息': '41', '燃气用户登记信息': '49', '入伍人员信息': '42', '社保单位信息': '99', '社保人员基本信息': '99', '上访人员信息': '45', '社会福利信息': '42', '社会服务信息': '99', '社会救助信息': '42', '手机短信信息': '52', '手机号段归属地信息': '52', '手机通话记录': '52', '手机通讯录信息': '52', '社区人员登记信息': '44', '社区网格化信息': '44', '收容人员信息': '42', '税务登记信息': '99', '事业单位人员信息': '41', '失业人员信息': '99', '收养信息': '42', '体检信息': '39', '通讯基站信息': '52', '体育健身单位登记信息': '99', '微博登记信息': '53', '网购交易信息': '53', '文化娱乐单位登记信息': '99', '物流信息': '35', '网络舆情信息': '53', '微信聊天信息': '53', '微信联系人信息': '53', '微信通讯录信息': '53', '微信用户信息': '53', '网游登记信息': '53', '修车信息': '37', '新浪微博信息': '53', '新农合医保人员信息': '99', '小区业主车辆登记信息': '44', '学生登记信息': '38', '新生儿出生信息': '39', '学校教员职工信息': '99', '信用社储户信息': '31', '医保刷卡记录信息': '39', '优抚安置信息': '99', '医疗机构患者信息': '39', '养老敬老信息': '42', '应聘登记信息': '53', '有线电视用户登记信息': '51', '银行交易信息': '31', '银行开户信息': '31', '邮政信息': '99', '住房公积金贷款信息': '30', '住房公积金还款信息': '30', '住房公积金信息': '30', '自来水缴费信息': '47', '自来水用户登记信息': '47', '住院登记信息': '39', '志愿服务信息': '42'}
def get_postday(offset):
    today = datetime.datetime.now()
    # 计算偏移量
    days = datetime.timedelta(days=offset)
    # 获取想要的日期的时间
    postday = (today + days).strftime('%Y%m%d')
    return int(postday)

def open_xlsx(file):
    try:
        # 打开excle文件，获取工作簿对象
        workbook = openpyxl.load_workbook(file)
        return workbook
    except Exception as e:
        print(file + '打开异常：' + e)

workbook = open_xlsx(excel_path)
# mark_date = get_postday(0)   #上报某天的数据入库：-1代表昨天，0代表当天
mark_date = get_postday(0)   #上报某天的数据入库：-1代表昨天，0代表当天
all_sheets = workbook.sheetnames   #所有sheet页存入列表

# 1.分县局/警种code 2.一级分类 3.上传文件名 4.101表名 5.最终入库数据量 6.单位类型代码 整型 7.IP段限制：默认1 整型 8.更新时间(当前时间)
# 9.创建人:万方 # 10.创建时间(当前时间) 11.上报时间(upload_datetime) 12.是否为内部数据：默认1 整型 13.上传方式 整型 14.数据质量 15.上报数据量
stat_list = []
for sheet_idx in range(2,len(all_sheets)):   #递归获取每一个sheet页(不包括：汇聚2020统计、汇聚2020统计、晾晒表)
    cur_sheet = workbook.get_sheet_by_name(all_sheets[sheet_idx])  # 获取当前sheet页对象

    depart_name = all_sheets[sheet_idx]  # 分县局/警钟名称
    print(depart_name)
    if depart_name in depart_dict.keys():
        depart_code = str(depart_dict[depart_name])
    else:depart_code = '99'

    # 单位类型。0：分县局；1：警种； 2：非考核单位，如：反恐
    if cur_sheet.cell(row=1, column=1).value == '序号0':
        depart_type_code = 0
    elif cur_sheet.cell(row=1, column=1).value == '序号1' and depart_name in nonExam_list:
        depart_type_code = 2
    else:
        depart_type_code = 1

    for row_idx in range(2, cur_sheet.max_row + 1):  # 递归获取当前sheet页每一行
        if cur_sheet.cell(row=row_idx, column=8).value and type(cur_sheet.cell(row = row_idx,column = 8).value) is int:
            upload_date = cur_sheet.cell(row=row_idx, column=8).value
            # liruoyu
            if upload_date == mark_date:
                if cur_sheet.cell(row=row_idx, column=6).value and cur_sheet.cell(row=row_idx, column=6).value.strip() and cur_sheet.cell(row=row_idx, column=6).value.strip() in tables_dict.keys():
                    table_cnname = cur_sheet.cell(row=row_idx, column=6).value.strip()
                    table_type_code = tables_dict[table_cnname]

                else:
                    table_cnname=None
                    table_type_code = '99'

                if cur_sheet.cell(row=row_idx, column=7).value and cur_sheet.cell(row=row_idx, column=7).value == 1:
                    upload_method = 1
                    if cur_sheet.cell(row=row_idx, column=2).value:
                        file_name = cur_sheet.cell(row=row_idx, column=2).value
                    else:
                        file_name = None
                        print(depart_name + "：FTP上传的第2列文件名字段错误")
                elif cur_sheet.cell(row=row_idx, column=7).value and cur_sheet.cell(row=row_idx, column=7).value == 2:
                    upload_method = 2
                    file_name = None
                else:
                    upload_method = 2
                    file_name = None
                    print(depart_name + "：第7列上传方式字段错误")

                if cur_sheet.cell(row=row_idx, column=9).value and cur_sheet.cell(row=row_idx, column=9).value.strip():
                    file_desc = cur_sheet.cell(row=row_idx, column=9).value.strip()
                else:file_desc = '--'

                if cur_sheet.cell(row=row_idx, column=3).value and type(cur_sheet.cell(row=row_idx, column=7).value) is int:
                    table_data_count = cur_sheet.cell(row=row_idx, column=3).value

                cur_datetime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                upload_datetime = datetime.datetime.strptime(str(upload_date), '%Y%m%d')


                stat_list.append((depart_code,table_type_code, table_data_count, depart_type_code, 1, cur_datetime, '万方', cur_datetime, upload_datetime, 1, upload_method, file_desc, table_data_count,file_name, table_cnname))
                # print(table_data_count)

        else:print(depart_name + "：第8列时间字段为空或非整型")
# print(stat_list)


conn = pymysql.connect(
        host=mysql_host,
        port=3306,
        # user='alpview',
        # password='123456',
        user='root',
        password='root',
        # db='cms',
        db='drying',
        charset='utf8'
)
cursor = conn.cursor()
sql = """INSERT INTO t_collection_situation (name,file_name, content_amount, type, state, update_time,create_by,create_time,report_time,is_inner,upload_method,data_quality,report_amount,upload_file,table_name) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
# sql = """INSERT INTO stat_test (name,file_name, content_amount, type, state, update_time,create_by,create_time,report_time,is_inner,upload_method,data_quality,report_amount,upload_file,table_name) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
try:
    print('插入MySQL开始')
    cursor.executemany(sql, stat_list)
    conn.commit()
    print('插入MySQL完成')
except Exception as e:
    print("插入MySQL异常:" + e)
    conn.rollback()
finally:
    cursor.close()
    conn.close()

workbook.close()   #关闭工作薄
print("脚本结束")