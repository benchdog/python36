# -*- coding: utf-8 -*-

# import xlwt  #只能写Excel
# import xlrd   #只能读Excel
# import xlutils  #修改Excel，在原来的基础上修改
import datetime
import openpyxl
from openpyxl.styles import Border,Side,colors,PatternFill,Alignment  #边框、填充颜色、对齐方式

# tables_dict = {'SJZS_BXLPXX_TMP': ['ZHE-22-443600002', '保险理赔信息', '保险', ], 'SJZS_BXXX_TMP': ['ZHE-22-413600001', '保险信息', '保险'], 'SJZS_BYSXX_TMP': ['ZPF-22-113600002', '毕业生信息', '教育'], 'SJZS_CJRDJXX_TMP': ['ZTB-22-113600015', '残疾人登记信息', '民政'], 'SJZS_CPJFXGRYXX_TMP': ['unknown', '产品纠纷相关人员信息', 'unknown'], 'SJZS_CSJKXX_TMP': ['ZTB-22-113600011', '慈善捐款信息', '民政'], 'SJZS_CZCDJXX_TMP': ['QJA-22-113600006', '出租车登记信息', '公共交通'], 'SJZS_CZCGJXX_TMP': ['QJA-22-143600008', '出租车轨迹信息', '公共交通'], 'SJZS_CZFWDJXX_TMP': ['ZEE-11-414200017', '出租房屋登记信息', '网络'], 'SJZS_CZYKT_TMP': ['unknown', '财政一卡通', 'unknown'], 'SJZS_DBJTXX_TMP': ['unknown', '低保家庭信息', 'unknown'], 'SJZS_DBRYXX_TMP': ['unknown', '低保人员信息', 'unknown'], 'SJZS_DDXX_TMP': ['ZKC-22-143200003', '典当信息', '服务业'], 'SJZS_DLYHDJXX_TMP': ['QAB-22-113600001', '电力用户登记信息', '供电'], 'SJZS_DWCYRYXX_TMP': ['ZSC-22-113600003', '单位从业人员信息', '劳动就业'], 'SJZS_DYXX_TMP': ['ZKB-22-113600002', '导游信息', '旅游'], 'SJZS_DZYXDJXX_TMP': ['ZEE-11-414200011', '电子邮箱登记信息', '网络'], 'SJZS_FLYZDXXX_TMP': ['ZNE-22-113200001', '法律援助对象信息', '司法'], 'SJZS_FWDJXX_TMP': ['ZFE-22-213200001', '房屋登记信息', '房地产'], 'SJZS_FWJYXX_TMP': ['ZFE-22-243200002', '房屋交易信息', '房地产'], 'SJZS_FXXX_TMP': ['unknown', '抚恤信息', 'unknown'], 'SJZS_GDLDRYXX_TMP': ['ZSC-22-113600004', '工地流动人员信息', '劳动就业'], 'SJZS_GGZXCKXX_TMP': ['unknown', '公共自行车卡信息', 'unknown'], 'SJZS_GJCYRYXX_TMP': ['QJA-22-113600001', '公交从业人员信息', '公共交通'], 'SJZS_GJYKTCZXX_TMP': ['QJA-22-143600003', '公交一卡通充值信息', '公共交通'], 'SJZS_GJYKTJYXX_TMP': ['QJA-22-143600004', '公交一卡通交易信息', '公共交通'], 'SJZS_GJYKTTKXX_TMP': ['QJA-22-143600005', '公交一卡通退卡信息', '公共交通'], 'SJZS_GJYKTYHXX_TMP': ['QJA-22-113600002', '公交一卡通用户信息', '公共交通'], 'SJZS_GKHCRYXX_TMP': ['unknown', '高考核查人员信息', 'unknown'], 'SJZS_GLKYGPXX_TMP': ['QJB-22-143600002', '公路客运购票信息', '公路'], 'SJZS_GNXX_TMP': ['QAD-22-113600001', '供暖信息', '供暖'], 'SJZS_GSGLCLJZSFTXXX_TMP': ['QJB-22-243200004', '高速公路车辆计重收费通行信息', '公路'], 'SJZS_GSGLCLTXXX_TMP': ['unknown', '高速公路车辆通行信息', 'unknown'], 'SJZS_GSRYXX_TMP': ['unknown', '工伤人员信息', 'unknown'], 'SJZS_GWRYXX_TMP': ['ZSC-22-113100006', '公务人员信息', '劳动就业'], 'SJZS_GWWZDJXX_TMP': ['ZEE-11-414200012', '购物网站登记信息', '网络'], 'SJZS_GZFRZXX_TMP': ['ZFE-22-243200003', '公租房入住信息', '房地产'], 'SJZS_GZKXX_TMP': ['ZPF-22-113600004', '高自考信息', '教育'], 'SJZS_HDXXZEB_TMP': ['ZEB-23-464200002', '话单信息(ZEB)', '通信'], 'SJZS_HHXX_TMP': ['ZTB-22-113600019', '火化信息', '民政'], 'SJZS_HYKDJXX_TMP': ['ZJA-22-413600001', '会员卡登记信息', '商业、贸易'], 'SJZS_HYKSKXX_TMP': ['ZJA-22-443600002', '会员卡刷卡信息', '商业、贸易'], 'SJZS_HYZKXX_TMP': ['ZTB-22-113600020', '婚姻状况信息', '民政'], 'SJZS_JDZYGB_TMP': ['unknown', '军队转业干部', 'unknown'], 'SJZS_JDZZZYRYXX_TMP': ['unknown', '军队自主择业人员信息', 'unknown'], 'SJZS_JHDJXX_TMP': ['ZTB-22-113600001', '结婚登记信息', '民政'], 'SJZS_JHSYFWDXDJXX_TMP': ['ZQJ-22-113600001', '计划生育服务对象登记信息', '计划生育'], 'SJZS_JSZYKXX_TMP': ['unknown', '驾驶证约考信息', 'unknown'], 'SJZS_JTYSWFAJXX_TMP': ['ZJD-22-313400003', '交通运输违法案件信息', '物流、仓储'], 'SJZS_JXYCLJJXX_TMP': ['unknown', '机修业车辆交接信息', 'unknown'], 'SJZS_JXYESCCLXX_TMP': ['ZKC-22-213200006', '机修业/二手车车辆信息', '服务业'], 'SJZS_JXYESCCYRYXX_TMP': ['ZKC-22-113200009', '机修业/二手车从业人员信息', '服务业'], 'SJZS_JYZHDJXX_TMP': ['ZEE-11-414200009', '交友征婚登记信息', '网络'], 'SJZS_JZFWDJXX_TMP': ['ZEE-11-414200013', '家政服务登记信息', '网络'], 'SJZS_JZXX_TMP': ['ZTB-22-113700012', '救助信息', '民政'], 'SJZS_KDXX_TMP': ['ZJD-22-113600002', '快递信息', '物流、仓储'], 'SJZS_KDYHDJXX_TMP': ['ZEE-11-113200001', '宽带用户登记信息', '网络'], 'SJZS_KJXX_TMP': ['unknown', '快件信息', 'unknown'], 'SJZS_KYCLDJXX_TMP': ['QJB-22-213600001', '客运车辆登记信息', '公路'], 'SJZS_LHDJXX_TMP': ['ZTB-22-113600002', '离婚登记信息', '民政'], 'SJZS_LTDJXX_TMP': ['ZEE-11-414200015', '论坛登记信息', '网络'], 'SJZS_LWYGSCDJXX_TMP': ['unknown', '劳务用工市场登记信息', 'unknown'], 'SJZS_LXSJDLKXX_TMP': ['ZKB-22-113600001', '旅行社接待旅客信息', '旅游'], 'SJZS_MZDJXX_TMP': ['ZQH-22-113600003', '门诊登记信息', '医疗保健'], 'SJZS_QCZLXX_TMP': ['ZKC-22-213600001', '汽车租赁信息', '服务业'], 'SJZS_QGLGVIPHY_TMP': ['unknown', '全国旅馆VIP会员', 'unknown'], 'SJZS_QQDJXX_TMP': ['ZEE-11-414200002', 'QQ登记信息', '网络'], 'SJZS_QQHYXX_TMP': ['ZEE-11-414200003', 'QQ好友信息', '网络'], 'SJZS_QQQXX_TMP': ['ZEE-11-414200004', 'QQ群信息', '网络'], 'SJZS_QZFWDJXX_TMP': ['ZEE-11-414200018', '求租房屋登记信息', '网络'], 'SJZS_RCSCDJXX_TMP': ['ZSC-22-113600001', '人才市场登记信息', '劳动就业'], 'SJZS_RQYHDJXX_TMP': ['QAC-22-113600001', '燃气用户登记信息', '供气'], 'SJZS_RWRYXX_TMP': ['ZTB-22-113600005', '入伍人员信息', '民政'], 'SJZS_SBDWXX_TMP': ['unknown', '社保单位信息', 'unknown'], 'SJZS_SBRYJBXX_TMP': ['unknown', '社保人员基本信息', 'unknown'], 'SJZS_SFRYXX_TMP': ['ZUF-22-113200001', '上访人员信息', '信访'], 'SJZS_SHFLXX_TMP': ['ZTB-22-113700018', '社会福利信息', '民政'], 'SJZS_SHFWXX_TMP': ['unknown', '社会服务信息', 'unknown'], 'SJZS_SHJZXX_TMP': ['ZTB-22-113700017', '社会救助信息', '民政'], 'SJZS_SJDXXX_TMP': ['ZEB-23-463200004', '手机短信信息', '通信'], 'SJZS_SJHDGSDXX_TMP': ['ZEB-23-413200007', '手机号段归属地信息', '通信'], 'SJZS_SJTHJL_TMP': ['ZEB-23-463200005', '手机通话记录', '通信'], 'SJZS_SJTXLXX_TMP': ['ZEB-23-413200003', '手机通讯录信息', '通信'], 'SJZS_SQRYDJXX_TMP': ['ZTC-22-113600001', '社区人员登记信息', '社区管理'], 'SJZS_SQWGHXX_TMP': ['ZTC-22-113600004', '社区网格化信息', '社区管理'], 'SJZS_SRRYXX_TMP': ['ZTB-22-113100013', '收容人员信息', '民政'], 'SJZS_SWDJXX_TMP': ['unknown', '税务登记信息', 'unknown'], 'SJZS_SYDWRYXX_TMP': ['ZSC-22-113600005', '事业单位人员信息', '劳动就业'], 'SJZS_SYRYXX_TMP': ['unknown', '失业人员信息', 'unknown'], 'SJZS_SYXX_TMP': ['ZTB-22-113600003', '收养信息', '民政'], 'SJZS_TJXX_TMP': ['ZQH-22-113600006', '体检信息', '医疗保健'], 'SJZS_TXJZXX_TMP': ['ZEB-23-413200006', '通讯基站信息', '通信'], 'SJZS_TYJSDWDJXX_TMP': ['unknown', '体育健身单位登记信息', 'unknown'], 'SJZS_WBDJXX_TMP': ['ZEE-11-414200016', '微博登记信息', '网络'], 'SJZS_WGJYXX_TMP': ['ZEE-11-414200006', '网购交易信息', '网络'], 'SJZS_WHYLDWDJXX_TMP': ['unknown', '文化娱乐单位登记信息', 'unknown'], 'SJZS_WLXX_TMP': ['ZJD-22-213600001', '物流信息', '物流、仓储'], 'SJZS_WLYQXX_TMP': ['ZEE-11-414200008', '网络舆情信息', '网络'], 'SJZS_WXLTXX_TMP': ['ZEE-11-444200022', '微信聊天信息', '网络'], 'SJZS_WXLXRXX_TMP': ['ZEE-11-434200020', '微信联系人信息', '网络'], 'SJZS_WXTXLXX_TMP': ['ZEE-11-434200021', '微信通讯录信息', '网络'], 'SJZS_WXYHXX_TMP': ['ZEE-11-414200019', '微信用户信息', '网络'], 'SJZS_WYDJXX_TMP': ['ZEE-11-414200014', '网游登记信息', '网络'], 'SJZS_XCXX_TMP': ['ZKC-22-243200007', '修车信息', '服务业'], 'SJZS_XLWBXX_TMP': ['ZEE-11-444200023', '新浪微博信息', '网络'], 'SJZS_XNHYBRYXX_TMP': ['unknown', '新农合医保人员信息', 'unknown'], 'SJZS_XQYZCLXX_TMP': ['ZTC-22-213600003', '小区业主车辆信息', '社区管理'], 'SJZS_XSDJXX_TMP': ['ZPF-22-113600001', '学生登记信息', '教育'], 'SJZS_XSECSXX_TMP': ['ZQH-22-113600005', '新生儿出生信息', '医疗保健'], 'SJZS_XXJYZGXX_TMP': ['unknown', '学校教员职工信息', 'unknown'], 'SJZS_XYSCHXX_TMP': ['ZHD-22-113600007', '信用社储户信息', '金融'], 'SJZS_YBSKJLXX_TMP': ['ZQH-22-143600002', '医保刷卡记录信息', '医疗保健'], 'SJZS_YFAZXX_TMP': ['unknown', '优抚安置信息', 'unknown'], 'SJZS_YLJGHZXX_TMP': ['ZQH-22-113600001', '医疗机构患者信息', '医疗保健'], 'SJZS_YLJLXX_TMP': ['ZTB-22-113600004', '养老敬老信息', '民政'], 'SJZS_YPDJXX_TMP': ['ZEE-11-414200010', '应聘登记信息', '网络'], 'SJZS_YXDSYHDJXX_TMP': ['ZGF-22-113600001', '有线电视用户登记信息', '广播、电影、电视'], 'SJZS_YXJYXX_TMP': ['ZHD-22-413600006', '银行交易信息', '金融'], 'SJZS_YXKHXX_TMP': ['ZHD-22-413600005', '银行开户信息', '金融'], 'SJZS_YZXX_TMP': ['unknown', '邮政信息', 'unknown'], 'SJZS_ZFGJJDKXX_TMP': ['ZFE-22-243200005', '住房公积金贷款信息', '房地产'], 'SJZS_ZFGJJHKXX_TMP': ['ZFE-22-243200006', '住房公积金还款信息', '房地产'], 'SJZS_ZFGJJXX_TMP': ['ZFE-22-213200004', '住房公积金信息', '房地产'], 'SJZS_ZLSJFXX_TMP': ['QAA-22-443600002', '自来水缴费信息', '供水'], 'SJZS_ZLSYHDJXX_TMP': ['QAA-22-113600001', '自来水用户登记信息', '供水'], 'SJZS_ZYDJXX_TMP': ['ZQH-22-113600004', '住院登记信息', '医疗保健'], 'SJZS_ZYFWXX_TMP': ['ZTB-22-113600014', '志愿服务信息', '民政']}
tables_dict = {'保险理赔信息': '33', '保险信息': '33', '毕业生信息': '38', '残疾人登记信息': '42', '产品纠纷相关人员信息': '99', '慈善捐款信息': '42', '出租车登记信息': '55', '出租车轨迹信息': '55', '出租房屋登记信息': '53', '财政一卡通': '99', '低保家庭信息': '99', '低保人员信息': '99', '典当信息': '37', '电力用户登记信息': '48', '单位从业人员信息': '41', '导游信息': '36', '电子邮箱登记信息': '53', '法律援助对象信息': '54', '房屋登记信息': '30', '房屋交易信息': '30', '抚恤信息': '99', '工地流动人员信息': '41', '公共自行车卡信息': '99', '公交从业人员信息': '55', '公交一卡通充值信息': '55', '公交一卡通交易信息': '55', '公交一卡通退卡信息': '55', '公交一卡通用户信息': '55', '高考核查人员信息': '99', '公路客运购票信息': '56', '供暖信息': '50', '高速公路车辆计重收费通行信息': '56', '高速公路车辆通行信息': '99', '工伤人员信息': '99', '公务人员信息': '41', '购物网站登记信息': '53', '公租房入住信息': '30', '高自考信息': '38', '话单流水信息': '52', '火化信息': '42', '会员卡登记信息': '34', '会员卡刷卡信息': '34', '婚姻状况信息': '42', '军队转业干部': '99', '军队自主择业人员信息': '99', '结婚登记信息': '42', '计划生育服务对象登记信息': '40', '驾驶证约考信息': '99', '交通运输违法案件信息': '35', '机修业车辆交接信息': '99', '机修业/二手车车辆信息': '37', '机修业/二手车从业人员信息': '37', '交友征婚登记信息': '53', '家政服务登记信息': '53', '救助信息': '42', '快递信息': '35', '宽带上网用户登记信息': '53', '客运车辆登记信息': '56', '离婚登记信息': '42', '论坛登记信息': '53', '劳动用工市场登记信息': '41', '旅行社接待旅客信息': '36', '门诊登记信息': '39', '汽车租赁信息': '37', '全国旅馆VIP会员': '99', 'QQ登记信息': '53', 'QQ好友信息': '53', 'QQ群信息': '53', '求租房屋登记信息': '53', '人才市场登记信息': '41', '燃气用户登记信息': '49', '入伍人员信息': '42', '社保单位信息': '99', '社保人员基本信息': '99', '上访人员信息': '45', '社会福利信息': '42', '社会服务信息': '99', '社会救助信息': '42', '手机短信信息': '52', '手机号段归属地信息': '52', '手机通话记录': '52', '手机通讯录信息': '52', '社区人员登记信息': '44', '社区网格化信息': '44', '收容人员信息': '42', '税务登记信息': '99', '事业单位人员信息': '41', '失业人员信息': '99', '收养信息': '42', '体检信息': '39', '通讯基站信息': '52', '体育健身单位登记信息': '99', '微博登记信息': '53', '网购交易信息': '53', '文化娱乐单位登记信息': '99', '物流信息': '35', '网络舆情信息': '53', '微信聊天信息': '53', '微信联系人信息': '53', '微信通讯录信息': '53', '微信用户信息': '53', '网游登记信息': '53', '修车信息': '37', '新浪微博信息': '53', '新农合医保人员信息': '99', '小区业主车辆登记信息': '44', '学生登记信息': '38', '新生儿出生信息': '39', '学校教员职工信息': '99', '信用社储户信息': '31', '医保刷卡记录信息': '39', '优抚安置信息': '99', '医疗机构患者信息': '39', '养老敬老信息': '42', '应聘登记信息': '53', '有线电视用户登记信息': '51', '银行交易信息': '31', '银行开户信息': '31', '邮政信息': '99', '住房公积金贷款信息': '30', '住房公积金还款信息': '30', '住房公积金信息': '30', '自来水缴费信息': '47', '自来水用户登记信息': '47', '住院登记信息': '39', '志愿服务信息': '42'}
excel_path= r'D:\guohui\汇聚2020\汇聚2020统计.xlsx'
# excel_path= r'C:\Users\bench\Desktop\2020科通数据汇聚\汇聚2020统计.xlsx'

def open_xlsx(file):
    try:
        # 打开excle文件，获取工作簿对象
        workbook = openpyxl.load_workbook(file)
        return workbook
    except Exception as e:
        print(file + '打开异常：' + str(e))

def get_postday(offset):
    today = datetime.datetime.now()
    # 计算偏移量
    days = datetime.timedelta(days=offset)
    # 获取想要的日期的时间
    postday = (today + days).strftime('%Y%m%d')
    return int(postday)

#根据当天周几，来获取当天所需统计时间段 postdays
#周一：上周五00:00 - 本周一17:00； 周二 至 周三：当天17:00之前； 周四：上周五00:00 至 本周四18:00
postdays = []
weekday_int = datetime.datetime.now().isoweekday()
# weekday_int = 2
if weekday_int in [1]:
    postdays = [get_postday(0), get_postday(-1), get_postday(-2), get_postday(-3)]
    # postdays = [get_postday(0), get_postday(-1), get_postday(-2), get_postday(-3), get_postday(-4)]
elif weekday_int in [2,3,5,6,7]:
    postdays = [get_postday(0)]
    # postdays = [get_postday(-1)]
elif weekday_int == 4:
    postdays = [get_postday(0), get_postday(-1), get_postday(-2), get_postday(-3), get_postday(-4), get_postday(-5), get_postday(-6)]
    # postdays = [get_postday(-1),get_postday(-2),get_postday(-3),get_postday(-4),get_postday(-5),get_postday(-6),get_postday(-7)]
else:
    postdays = [get_postday(0)]
print('今天周' + str(weekday_int) + ',统计时间段：' + str(postdays))

workbook = open_xlsx(excel_path)
# workbook = openpyxl.load_workbook(excel_path)
all_sheets = workbook.sheetnames   #所有sheet页存入列表
#“汇聚2020统计”：删除旧旧sheet，备份旧sheet，并隐藏
if '汇聚2020统计_old' in all_sheets:
    workbook.remove(workbook['汇聚2020统计_old'])   #删除旧备份sheet
workbook['汇聚2020统计'].title = '汇聚2020统计_old'   #备份旧sheet
workbook['汇聚2020统计_old'].sheet_state = 'hidden'   #隐藏备份的旧sheet
workbook['汇聚2020统计_old'].index = 1   #修改备份的旧sheet索引位置为1

#创建新sheet并格式化
workbook.create_sheet('汇聚2020统计',0)   #新建sheet'汇聚2020统计',索引位置为0
sheet_stat = workbook.get_sheet_by_name('汇聚2020统计')

#todo 不能获取合并单元格的值
#todo 超链接
#todo 排序
# stat_dict = {0:[],1:[]}   #0代表区县局，1代表局属单位
stat_list = [{},{}]   #stat_list[0]代表区县局，stat_list[1]代表局属单位
for sheet_idx in range(2,len(all_sheets)):   #递归获取每一个sheet页(不包括：汇聚2020统计、汇聚2020统计、晾晒表)
    # print(sheet_idx)
    sheet_sum = 0   #上报数据总量
    sheet_tables = []   #上报数据种类数量
    sheet_count = 0   #上报数据种类名称
    sheet_sum_incre = 0
    sheet_sum_append = 0   #(仅局属单位)增补类数据总量
    sheet_tables_before = []
    sheet_tables_after = []
    sheet_tables_incre = []
    sheet_count_incre = 0
    sheet_name = all_sheets[sheet_idx]   #获取当前sheet页名称
    print(sheet_name)
    cur_sheet = workbook.get_sheet_by_name(all_sheets[sheet_idx])   #获取当前sheet页对象
    for row_idx in range(2,len(list(cur_sheet.rows)) + 1):   #递归获取当前sheet页每一行
        #统计历史数据种类
        if cur_sheet.cell(row=row_idx, column=6).value and cur_sheet.cell(row=row_idx, column=6).value.strip():
            table_cnname = cur_sheet.cell(row=row_idx, column=6).value.strip()
            if table_cnname not in sheet_tables:
                if table_cnname in tables_dict.keys():
                    sheet_tables.append(table_cnname)
                else:
                    upload_type = 'FTP' if str(cur_sheet.cell(row=row_idx, column=7).value) == '1' else '采集工具'
                    print('<' + sheet_name + '：' + table_cnname + '：' + upload_type + '>' + ' 不在127张表范围内!')

            # 分别统计当天之前和之后的数据种类
            if table_cnname in tables_dict.keys() and table_cnname not in sheet_tables_before and cur_sheet.cell(row=row_idx,column=8).value and int(cur_sheet.cell(row=row_idx, column=8).value) not in postdays:
                sheet_tables_before.append(table_cnname)
            if table_cnname in tables_dict.keys() and table_cnname not in sheet_tables_after and cur_sheet.cell(row=row_idx,column=8).value and int(cur_sheet.cell(row=row_idx, column=8).value) in postdays:
                sheet_tables_after.append(table_cnname)


        # 统计历史数据总量
        if cur_sheet.cell(row=1, column=1).value == '序号0':   #序号0代表区县公安局
            if cur_sheet.cell(row=row_idx, column=3).value and type(cur_sheet.cell(row=row_idx, column=3).value) is int:
                sheet_sum += cur_sheet.cell(row=row_idx, column=3).value
                # 统计数据增量
                if cur_sheet.cell(row=row_idx, column=8).value and int(cur_sheet.cell(row=row_idx, column=8).value) in postdays:
                    sheet_sum_incre += cur_sheet.cell(row=row_idx, column=3).value

        elif cur_sheet.cell(row=1, column=1).value == '序号1':   #序号1代表局属单位(警种)
            if cur_sheet.cell(row=row_idx, column=6).value and cur_sheet.cell(row=row_idx, column=6).value.strip():
                table_cnname = cur_sheet.cell(row=row_idx, column=6).value.strip()
                if table_cnname in tables_dict.keys() and cur_sheet.cell(row=row_idx, column=3).value and type(cur_sheet.cell(row=row_idx, column=3).value) is int:
                    sheet_sum += cur_sheet.cell(row=row_idx, column=3).value
                    # 统计数据增量
                    if cur_sheet.cell(row=row_idx, column=8).value and int(cur_sheet.cell(row=row_idx, column=8).value) in postdays:
                        sheet_sum_incre += cur_sheet.cell(row=row_idx, column=3).value
            if not cur_sheet.cell(row=row_idx, column=6).value or not cur_sheet.cell(row=row_idx, column=6).value.strip():
                # print(cur_sheet.cell(row=row_idx, column=3).value)
                sheet_sum_append += cur_sheet.cell(row=row_idx, column=3).value
        else:
            print(all_sheets[sheet_idx] + '1.该sheet页备注列标记错误')

    for table in sheet_tables_after:
        if table not in sheet_tables_before:
            sheet_tables_incre.append(table)

    # 统计历史数据种类总数量
    sheet_count = len(sheet_tables)
    # 统计数据种类增量
    sheet_count_incre = len(sheet_tables_incre)
    if cur_sheet.cell(row=1, column=1).value == '序号0':
        stat_list[0][all_sheets[sheet_idx]] = [sheet_sum, sheet_count, sheet_tables, sheet_sum_incre, sheet_count_incre, sheet_tables_incre]
    elif cur_sheet.cell(row=1, column=1).value == '序号1':
        stat_list[1][all_sheets[sheet_idx]] = [sheet_sum, sheet_count, sheet_tables, sheet_sum_incre, sheet_count_incre, sheet_tables_incre, sheet_sum_append]
    else:
        print(all_sheets[sheet_idx] + '2.该sheet页备注列标记错误')

# print(stat_list)

#取消合并单元格
#第一步：获取已合并单元格的位置信息
merge_list = sheet_stat.merged_cells
cr = []
for merge_area in merge_list:
    # 合并单元格的起始行坐标、终止行坐标。。。。，
    r1, r2, c1, c2 = merge_area.min_row, merge_area.max_row, merge_area.min_col, merge_area.max_col
    # 纵向合并单元格的位置信息提取出
    if r2 - r1 > 0:
        cr.append((r1, r2, c1, c2))
#第二步：拆分单元格
for r in cr:
    sheet_stat.unmerge_cells(start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])

#合并单元格
row_mark = 1
#根据增量列排序(降序)
for e in sorted(stat_list[0].items(), key = lambda kv:(kv[1][3], kv[0]),reverse=True):
    row_mark += 1
    sheet_stat['A' + str(row_mark)] = row_mark - 1
    sheet_stat['C' + str(row_mark)] = e[0]
    sheet_stat['D' + str(row_mark)] = e[1][0]
    sheet_stat['E' + str(row_mark)] = e[1][1]
    sheet_stat['F' + str(row_mark)] = str(e[1][2]).strip('[').strip(']').replace("'",'')
    sheet_stat['G' + str(row_mark)] = e[1][3]
    sheet_stat['H' + str(row_mark)] = e[1][4]
    sheet_stat['I' + str(row_mark)] = str(e[1][5]).strip('[').strip(']').replace("'",'')
sheet_stat['B2'] = '分县区局'
sheet_stat.merge_cells('B2:' + 'B' + str(row_mark))

row_mark_mid = row_mark + 1
#根据增量列排序(降序)
for e in sorted(stat_list[1].items(), key = lambda kv:(kv[1][3], kv[0]),reverse=True):
    row_mark += 1
    sheet_stat['A' + str(row_mark)] = row_mark - 1
    sheet_stat['C' + str(row_mark)] = e[0]
    sheet_stat['D' + str(row_mark)] = e[1][0]
    sheet_stat['E' + str(row_mark)] = e[1][1]
    sheet_stat['F' + str(row_mark)] = str(e[1][2]).strip('[').strip(']').replace("'",'')
    sheet_stat['G' + str(row_mark)] = e[1][3]
    sheet_stat['H' + str(row_mark)] = e[1][4]
    sheet_stat['I' + str(row_mark)] = str(e[1][5]).strip('[').strip(']').replace("'",'')
    sheet_stat['J' + str(row_mark)] = e[1][6]
sheet_stat['B' + str(row_mark_mid)] = '局属单位'
sheet_stat.merge_cells('B' + str(row_mark_mid) + ':B' + str(row_mark))

#边框
border = Border(left=Side(border_style='thin',color=colors.BLACK),right=Side(border_style='thin',color=colors.BLACK),top=Side(border_style='thin',color=colors.BLACK),bottom=Side(border_style='thin',color=colors.BLACK))
#对齐方式
align = Alignment(horizontal='left',vertical='center',wrap_text=True)
#填充颜色
fill = PatternFill("solid", fgColor="FFFF00")
for r in range(1,row_mark + 1):
    sheet_stat['A' + str(r)].border = border
    sheet_stat['B' + str(r)].border = border
    sheet_stat['C' + str(r)].border = border
    sheet_stat['D' + str(r)].border = border
    sheet_stat['E' + str(r)].border = border
    sheet_stat['F' + str(r)].border = border
    sheet_stat['G' + str(r)].border = border
    sheet_stat['H' + str(r)].border = border
    sheet_stat['I' + str(r)].border = border
    sheet_stat['J' + str(r)].border = border

    sheet_stat['A' + str(r)].alignment = align
    sheet_stat['B' + str(r)].alignment = align
    sheet_stat['C' + str(r)].alignment = align
    sheet_stat['D' + str(r)].alignment = align
    sheet_stat['E' + str(r)].alignment = align
    sheet_stat['F' + str(r)].alignment = align
    sheet_stat['G' + str(r)].alignment = align
    sheet_stat['H' + str(r)].alignment = align
    sheet_stat['I' + str(r)].alignment = align
    sheet_stat['J' + str(r)].alignment = align

sheet_stat['A1'].fill = fill
sheet_stat['B1'].fill = fill
sheet_stat['C1'].fill = fill
sheet_stat['D1'].fill = fill
sheet_stat['E1'].fill = fill
sheet_stat['F1'].fill = fill
sheet_stat['G1'].fill = fill
sheet_stat['H1'].fill = fill
sheet_stat['I1'].fill = fill
sheet_stat['J1'].fill = fill

#设置首行表头
sheet_stat['A1'] = '序号'
sheet_stat['B1'] = '机构类别'
sheet_stat['C1'] = '机构名称'
sheet_stat['D1'] = '总量'
sheet_stat['E1'] = '种类'
sheet_stat['F1'] = '种类名称'
sheet_stat['G1'] = '增量'
sheet_stat['H1'] = '增量种类'
sheet_stat['I1'] = '增量种类名称'
sheet_stat['J1'] = '增补类总量'


workbook.save(excel_path)   #保存工作薄
workbook.close()   #关闭工作薄
print("统计完成")