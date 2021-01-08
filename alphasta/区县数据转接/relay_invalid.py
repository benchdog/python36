#coding: utf8
from docx import Document
from docx.oxml.ns import qn
from docx.shared import Pt,RGBColor
import pymysql
import time
import openpyxl
from openpyxl import Workbook
import shutil
import os
from ftplib import FTP

# docx文档每行内容
p1 = "    为落实全省公共安全视频智能化建设应用专项治理工作调度会精神，提升我市视频监控前端在线率， 促进我市视频智能化建设应用工作提档升级，按照中心领导要求，每日对各县（市、区）各类视频监控前端在线、数据质量情况进行通报。\n\n    一、视频监控前端在线率情况\n"
p2_civic = "    市内五区\n    长安：     %，桥西：    %，新华：    %，裕华：     %，高新：    %\n\n    二、统计全市智能前端数据质量情况\n    1.人脸智能前端无数据上传数量\n    市内五区\n    "
p2_county = "    无极：  藁城：  栾城：  鹿泉：  正定：  高邑： 新乐：  赞皇：  元氏：  \n    行唐：  循环化工：  井陉：   矿区：  平山：  晋州：  赵县：  灵寿：  深泽： \n\n    二、统计全市智能前端数据质量情况\n    1.人脸智能前端无数据上传数量\n    "
p4_civic = "\n    2.车辆智能前端无数据上传数量\n    市内五区\n    "
p4_county = "\n    2.车辆智能前端无数据上传数量\n    "
p6 = "\n\n    三、录像可用性\n    不可调用的有"
def docx_edit(type,p,path):
    try:
        docx = Document()
        docx.styles['Normal'].font.name = u'宋体'
        docx.styles['Normal']._element.rPr.rFonts.set(qn('w:eastAsia'), u'宋体')
        docx.styles['Normal'].font.size = Pt(10.5)
        docx.styles['Normal'].font.color.rgb = RGBColor(0,0,0)
        docx.add_paragraph(p)
        docx.add_page_break()
        docx.save(path)
        print(str(type) + '人脸/车辆无效设备统计已写入docx')
    except Exception as e:
        print('docx文档编辑异常：' + str(e))

#当天日期
today=time.strftime("%Y-%m-%d", time.localtime())
dir = r'C:\Users\bench\Desktop\ligh\万方\视频中心区县数据转接\无效设备历史统计'
# dir = '/home/ligh/relay_stat'
civic = '市区视频数据统计通报'
county = '县域视频数据统计通报'
civic_today = civic + '_' + str(today)
county_today = county + '_' + str(today)

#连接ftp
# try:
#     ftp = FTP()
#     ftp.connect("13.32.4.176", 21)  #第一个参数可以是ftp服务器的ip或者域名，第二个参数为ftp服务器的连接端口，默认为21
#     ftp.login('files', '')   #匿名登录直接使用ftp.login()
#     # ftp.cwd('县域视频数据统计通报'.encode('gbk').decode('iso-8859-1'))
# except Exception as e:
#     print('ftp连接异常：' + str(e))
#
# def ftp_upload(ftp, cwd, src, dst):
#     try:
#         # ftp.cwd(cwd.encode('gbk').decode('iso-8859-1'))
#         ftp.cwd(cwd)
#         fp = open(src, 'rb')
#         ftp.storbinary('STOR ' + dst.encode('gbk').decode('iso-8859-1'), fp, 1024)
#         print(str(dst) + '上传ftp成功')
#     except Exception as e:
#         print(str(dst) + '上传ftp失败：' + str(e))


def ftp_upload(cwd, src, dst):
    try:
        ftp = FTP()
        ftp.connect("13.32.4.176", 21)  #第一个参数可以是ftp服务器的ip或者域名，第二个参数为ftp服务器的连接端口，默认为21
        ftp.login('files', '')   #匿名登录直接使用ftp.login()
        # ftp.cwd('县域视频数据统计通报'.encode('gbk').decode('iso-8859-1'))
        ftp.cwd(cwd.encode('gbk').decode('iso-8859-1'))
        fp = open(src, 'rb')
        ftp.storbinary('STOR ' + dst.encode('gbk').decode('iso-8859-1'), fp, 1024)
        print(str(dst) + '上传ftp成功')
    except Exception as e:
        print('上传ftp失败：' + str(e))
    finally:
        ftp.close()   #关闭ftp

conn = pymysql.connect(
        host='13.32.4.170',
        # host='192.168.23.112',
        port=3306,
        user='root',
        password='wanfang@2001',
        db='db1400',
        # db='ligh',
        charset='utf8'
)
def mysql_select(sql):
    res_select = ()
    try:
        # print('查询...')
        cursor = conn.cursor()
        cursor.execute(sql)
        res_select = cursor.fetchall()
    except Exception as e:
        print("查询异常:" + str(e))
    finally:
        cursor.close()
        return res_select

def open_xlsx(file):
    try:
        # 打开excle文件，获取工作簿对象
        workbook = openpyxl.load_workbook(file)
        return workbook
    except Exception as e:
        print(file + ' 打开异常：' + str(e))

def copy_file(src_file,dst_file):
    if not os.path.exists(dst_file):
        shutil.copy(src_file, dst_file)
        print(str(dst_file) + "创建成功！")

# copy_file(os.path.join(dir,civic + '.xlsx'), os.path.join(dir, civic_today + '.xlsx' ))
# copy_file(os.path.join(dir,civic + '.docx'), os.path.join(dir, civic_today + '.docx' ))
# copy_file(os.path.join(dir,county + '.xlsx'), os.path.join(dir, county_today + '.xlsx' ))
# copy_file(os.path.join(dir,county + '.docx'), os.path.join(dir, county_today + '.docx' ))


# workbook_civic = open_xlsx(os.path.join(dir, civic_today + '.xlsx' ))
# workbook_county = open_xlsx(os.path.join(dir, county_today + '.xlsx' ))
#
# sheet_civic_collect = workbook_civic.get_sheet_by_name('市区设备汇总')
# sheet_civic_collect.delete_rows(2,50)
# sheet_civic_detail = workbook_civic.get_sheet_by_name('市区设备明细')
# sheet_civic_detail.delete_rows(2,5000)
#
# sheet_county_collect = workbook_county.get_sheet_by_name('县域设备汇总')
# sheet_county_collect.delete_rows(2,50)
# sheet_county_detail = workbook_county.get_sheet_by_name('县域设备明细')
# sheet_county_detail.delete_rows(2,5000)

workbook_civic = Workbook()  # 创建工作簿
sheet_civic_collect = workbook_civic.create_sheet("市区设备汇总")  # 创建mysheet表
sheet_civic_detail = workbook_civic.create_sheet("市区设备明细")
workbook_county = Workbook()  # 创建工作簿
sheet_county_collect = workbook_county.create_sheet("县域设备汇总")  # 创建mysheet表
sheet_county_detail = workbook_county.create_sheet("县域设备明细")

workbook_civic.remove(workbook_civic['Sheet'])
workbook_county.remove(workbook_county['Sheet'])

sheet_civic_collect['A1'] = '序号'
sheet_civic_collect['B1'] = '区县厂商'
sheet_civic_collect['C1'] = '人脸'
sheet_civic_collect['D1'] = '车辆'
sheet_civic_collect['E1'] = '未知'
sheet_civic_detail['A1'] = '序号'
sheet_civic_detail['B1'] = '区县厂商'
sheet_civic_detail['C1'] = '设备类型'
sheet_civic_detail['D1'] = '设备ID'
sheet_civic_detail['E1'] = '设备名称'
sheet_civic_detail['F1'] = '设备IP'

sheet_county_collect['A1'] = '序号'
sheet_county_collect['B1'] = '区县厂商'
sheet_county_collect['C1'] = '人脸'
sheet_county_collect['D1'] = '车辆'
sheet_county_collect['E1'] = '未知'
sheet_county_detail['A1'] = '序号'
sheet_county_detail['B1'] = '区县厂商'
sheet_county_detail['C1'] = '设备类型'
sheet_county_detail['D1'] = '设备ID'
sheet_county_detail['E1'] = '设备名称'


#市内5区统计
sql_civic="select data_source,function_type,id,name,ip from ape_civic where id not in (select device_id from t_receive_data_log where date=CURDATE() AND type not in ('3','7'))"
res_civic=mysql_select(sql_civic)
# print(res_civic)
dict_civic={}
if res_civic:
    for e in res_civic:
        if e[0] not in dict_civic.keys():
            if e[1] == '人脸':
                dict_civic.update({e[0]:[[(e[2], e[3], e[4])], [], []]})
            elif e[1] == '车辆':
                dict_civic.update({e[0]: [[], [([e[2], e[3], e[4]])], []]})
            else:
                dict_civic.update({e[0]: [[], [], [([e[2], e[3], e[4]])]]})
        else:
            if e[1] == '人脸':
                dict_civic[e[0]][0].append(([e[2], e[3], e[4]]))
            elif e[1] == '车辆':
                dict_civic[e[0]][1].append(([e[2], e[3], e[4]]))
            else:
                dict_civic[e[0]][2].append(([e[2], e[3], e[4]]))
# print(dict_civic)


dict_civic_face = {}
dict_civic_motor = {}
p3_civic = ''
p5_civic = ''
for k,v in dict_civic.items():
    dict_civic_face[k] = len(v[0])
for i in sorted(dict_civic_face.items(), key=lambda d: d[1], reverse=False):
    p3_civic += (i[0] + '：' + str(i[1]) + ' ')

for k,v in dict_civic.items():
    dict_civic_motor[k] = len(v[1])
for i in sorted(dict_civic_motor.items(), key=lambda d: d[1], reverse=False):
    p5_civic += (i[0] + '：' + str(i[1]) + ' ')
print('市区人脸:\n',p3_civic.strip())
print('市区车辆:\n',p5_civic.strip())
docx_edit('市区', p1 + p2_civic + p3_civic + p4_civic + p5_civic + p6, os.path.join(dir, civic_today + '.docx' ))



l1 = 2
l2 = 2
for k,v in dict_civic.items():
    # print(k,len(v[0]),len(v[1]),len(v[2]))
    sheet_civic_collect['A' + str(l1)] = str(l1-1)
    sheet_civic_collect['B' + str(l1)] = k
    sheet_civic_collect['C' + str(l1)] = len(v[0])
    sheet_civic_collect['D' + str(l1)] = len(v[1])
    sheet_civic_collect['E' + str(l1)] = len(v[2])
    l1 += 1

    for i in range(len(v[0])):
        sheet_civic_detail['A' + str(l2)] = str(l2 - 1)
        sheet_civic_detail['B' + str(l2)] = k
        sheet_civic_detail['C' + str(l2)] = '人脸'
        sheet_civic_detail['D' + str(l2)] = v[0][i][0]
        sheet_civic_detail['E' + str(l2)] = v[0][i][1]
        sheet_civic_detail['F' + str(l2)] = v[0][i][2]
        l2 += 1

    for i in range(len(v[1])):
        sheet_civic_detail['A' + str(l2)] = str(l2 - 1)
        sheet_civic_detail['B' + str(l2)] = k
        sheet_civic_detail['C' + str(l2)] = '车辆'
        sheet_civic_detail['D' + str(l2)] = v[1][i][0]
        sheet_civic_detail['E' + str(l2)] = v[1][i][1]
        sheet_civic_detail['F' + str(l2)] = v[1][i][2]
        l2 += 1

    for i in range(len(v[2])):
        sheet_civic_detail['A' + str(l2)] = str(l2 - 1)
        sheet_civic_detail['B' + str(l2)] = k
        sheet_civic_detail['C' + str(l2)] = '未知'
        sheet_civic_detail['D' + str(l2)] = v[2][i][0]
        sheet_civic_detail['E' + str(l2)] = v[2][i][1]
        l2 += 1
print('市区人脸/车辆无效设备统计已写入excel\n<--------->')

#18县统计
sql_county= "SELECT t0.name,CASE t1.FUNCTION_type WHEN '1' THEN '车辆' when '2' THEN '人脸' ELSE '未知' END AS '设备类别',t1.ape_id,t1.name from ((SELECT deviceId,name from t_viid_system where id not in ('1','32','29','34','37') AND type=1)t0 LEFT JOIN (select data_source,FUNCTION_type,ape_id,name from ape where ape_id not in (SELECT DISTINCT device_Id from t_push_data_log WHERE date=CURDATE() and result =1 AND data_source='13010020205035164320' AND type in ('12','13')))t1 ON t0.deviceId = t1.data_source)"
res_county=mysql_select(sql_county)
# print(res_county)
dict_county={}
if res_county:
    for e in res_county:
        if e[0] not in dict_county.keys():
            if e[1] == '人脸':
                dict_county.update({e[0]:[[([e[2], e[3]])], [], []]})
            elif e[1] == '车辆':
                dict_county.update({e[0]: [[], [([e[2], e[3]])], []]})
            else:
                dict_county.update({e[0]: [[], [], [([e[2], e[3]])]]})
        else:
            if e[1] == '人脸':
                dict_county[e[0]][0].append(([e[2], e[3]]))
            elif e[1] == '车辆':
                dict_county[e[0]][1].append(([e[2], e[3]]))
            else:
                dict_county[e[0]][2].append(([e[2], e[3]]))
# print(dict_county)




for k,v in dict_civic.items():
    dict_civic_face[k] = len(v[0])
for i in sorted(dict_civic_face.items(), key=lambda d: d[1], reverse=False):
    p3_civic += (i[0] + '：' + str(i[1]) + ' ')

for k,v in dict_civic.items():
    dict_civic_motor[k] = len(v[1])
for i in sorted(dict_civic_motor.items(), key=lambda d: d[1], reverse=False):
    p5_civic += (i[0] + '：' + str(i[1]) + ' ')



dict_county_face = {}
dict_county_motor = {}
p3_county = ''
p5_county = ''
for k,v in dict_county.items():
    dict_county_face[k] = len(v[0])
for i in sorted(dict_county_face.items(), key=lambda d: d[1], reverse=False):
    p3_county += (i[0][:-2] + '：' + str(i[1]) + ' ')

for k,v in dict_county.items():
    dict_county_motor[k] = len(v[1])
for i in sorted(dict_county_motor.items(), key=lambda d: d[1], reverse=False):
    p5_county += (i[0][:-2] + '：' + str(i[1]) + ' ')
print('县域人脸:\n',p3_county.strip())
print('县域车辆:\n',p5_county.strip())
docx_edit('县域', p1 + p2_county + p3_county + p4_county + p5_county + p6, os.path.join(dir, county_today + '.docx' ))


l1 = 2
l2 = 2
len1 = 0
len2 = 0
len3 = 0
for k, v in dict_county.items():
    # print(k,len(v[0]),len(v[1]),len(v[2]))

    sheet_county_collect['A' + str(l1)] = str(l1 - 1)
    sheet_county_collect['B' + str(l1)] = k
    sheet_county_collect['C' + str(l1)] = len(v[0])
    sheet_county_collect['D' + str(l1)] = len(v[1])
    sheet_county_collect['E' + str(l1)] = len(v[2])
    l1 += 1

    #汇总行统计
    len1 += len(v[0])
    len2 += len(v[1])
    len3 += len(v[2])

    for i in range(len(v[0])):
        sheet_county_detail['A' + str(l2)] = str(l2 - 1)
        sheet_county_detail['B' + str(l2)] = k
        sheet_county_detail['C' + str(l2)] = '人脸'
        sheet_county_detail['D' + str(l2)] = v[0][i][0]
        sheet_county_detail['E' + str(l2)] = v[0][i][1]
        l2 += 1

    for i in range(len(v[1])):
        sheet_county_detail['A' + str(l2)] = str(l2 - 1)
        sheet_county_detail['B' + str(l2)] = k
        sheet_county_detail['C' + str(l2)] = '车辆'
        sheet_county_detail['D' + str(l2)] = v[1][i][0]
        sheet_county_detail['E' + str(l2)] = v[1][i][1]
        l2 += 1

    for i in range(len(v[2])):
        sheet_county_detail['A' + str(l2)] = str(l2 - 1)
        sheet_county_detail['B' + str(l2)] = k
        sheet_county_detail['C' + str(l2)] = '未知'
        sheet_county_detail['D' + str(l2)] = v[2][i][0]
        sheet_county_detail['E' + str(l2)] = v[2][i][1]
        l2 += 1

#汇总行统计写入excel
sheet_county_collect['A' + str(l1)] = str(l1 - 1)
sheet_county_collect['B' + str(l1)] = '汇总'
sheet_county_collect['C' + str(l1)] = len1
sheet_county_collect['D' + str(l1)] = len2
sheet_county_collect['E' + str(l1)] = len3

print('县域人脸/车辆无效设备统计已写入excel\n<--------->')

#保存、关闭两个excel文件
workbook_civic.save(os.path.join(dir, civic_today + '.xlsx' ))
workbook_civic.close()
workbook_county.save(os.path.join(dir, county_today + '.xlsx' ))
workbook_county.close()

#县域统计上传ftp
# '市区视频数据统计通报'.encode('gbk').decode('iso-8859-1')
ftp_upload(county, os.path.join(dir, county_today + '.docx'), county_today + '.docx')
# time.sleep(10)
ftp_upload(county, os.path.join(dir, county_today + '.xlsx'), county_today + '.xlsx')
#市区统计上传ftp
ftp_upload(civic, os.path.join(dir, civic_today + '.docx'), civic_today + '.docx')
ftp_upload(civic, os.path.join(dir, civic_today + '.xlsx'), civic_today + '.xlsx')
print('<--------->')

conn.close()   #关闭mysql连接
print('脚本完成!')