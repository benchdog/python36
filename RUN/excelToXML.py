import openpyxl
from xml.dom.minidom import Document

# 从工作薄中获取一个表单(sheet)对象
# sheets = wb.sheetnames
# print(type(sheets),sheets)

# def excelReader(excelPath):
#     # 打开excel文件,获取工作簿对象
#     wb = openpyxl.load_workbook(excelPath)
#             # print(sheet.title)
#             # col_range = sheet['A:I']


'''
对VIID20类数据的元数据转换成XML格式文档
'''
# 创建xml doc
def metaXMLWriter(excelPath, xmlPath, tagList):
    # 创建doc
    doc = Document()
    # doc.createComment('encoding="UTF-8"')
    # 创建DataSetFile根节点
    root = doc.createElement(tagList[0])
    root.setAttribute('ENName','VIID')
    root.setAttribute('CHName','视图库数据集文件')
    root.setAttribute('Description','')
    doc.appendChild(root)

    wb = openpyxl.load_workbook(excelPath)
    for sheet in wb:
        if sheet.title != '字典集':
            # 创建DataSet节点
            dataset = doc.createElement(tagList[1])
            dataset.setAttribute('CHName', sheet['A1'].value)
            dataset.setAttribute('ENName', sheet['B1'].value)
            dataset.setAttribute('Description', sheet['C1'].value)
            root.appendChild(dataset)

            row_range = sheet[3:sheet.max_row]
            # print(sheet.max_row)
            # field = doc.createElement(tagList[2])
            for row in row_range:
                # 创建Field节点
                field = doc.createElement(tagList[2])
                # print(type(row))
                # print(row[0].value)
                # for cell in row:
                # print(row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value)

                field.setAttribute('CHName', row[0].value)
                field.setAttribute('ENName', row[1].value)
                field.setAttribute('ValueType', row[2].value)
                field.setAttribute('ValueLength', str(row[3].value))
                field.setAttribute('BeNotNull', row[4].value)
                field.setAttribute('CodeSet', row[5].value)
                field.setAttribute('BeQueried', row[6].value)
                # field.setAttribute('Description', row[7].value)
                field.setAttribute('Description', row[7].value.replace('\n',';') if row[7].value else row[7].value)
                field.setAttribute('ValueDefault', row[8].value)
                dataset.appendChild(field)
    with open(xmlPath, 'w',encoding='utf8') as fw:
        fw.write(doc.toprettyxml(indent='\t'))
    print('元字段转换完成')

# ws = wb.active  # 当前活跃的表单
# print(ws)
# print(ws['A99']) # 获取A列的第一个对象
# print(ws['A99'].value)


'''
以下是对字典集的XML文件生成
'''

def dictXMLWriter(excelPath, xmlPath, tagList):
    # 创建doc
    doc = Document()
    # doc.createComment('encoding="UTF-8"')
    # 创建DataSetFile根节点
    root = doc.createElement(tagList[0])
    root.setAttribute('Name','GAT1400.3-2017')
    # root.setAttribute('CHName','视图库字段分类代码表')
    root.setAttribute('Description','视图库字段分类代码表')
    doc.appendChild(root)

    wb = openpyxl.load_workbook(excelPath)
    dictSheet = wb.get_sheet_by_name('字典集')
    row_range = dictSheet[2:dictSheet.max_row]
    for row in row_range:
        # print(row[2].value)
        if row[2].value == 0:
            # print(row[2].value)
            codeset = doc.createElement(tagList[1])
            codeset.setAttribute('CHName', row[0].value)
            codeset.setAttribute('CSID', row[1].value.split('#')[1])
            root.appendChild(codeset)

        else:
            item = doc.createElement(tagList[2])
            item.setAttribute('Code', str(row[3].value))
            item.setAttribute('CHName', row[4].value)
            item.setAttribute('Description', row[5].value)
            codeset.appendChild(item)

    with open(xmlPath, 'w',encoding='utf8') as fw:
        fw.write(doc.toprettyxml(indent='\t'))
    print('字典集转换完成')

if __name__ == '__main__':
    excelPath = 'E:\\gacp\\docs\\开发过程文档\\20类数据元数据整理.xlsx'
    # rootAttrDict = {'ENName':'VIID','CHName':'视图库数据集文件','Description':''}
    metaxmlPath = 'C:/Users/think/Desktop/VIID_Dataset.xml'
    dictxmlPath = 'C:/Users/think/Desktop/GAT1400.3-2017.xml'
    metaTagList = ['DataSetFile', 'DataSet', 'Field']
    dictTagList = ['CodeSetFile', 'CodeSet', 'Item']
    metaXMLWriter(excelPath, metaxmlPath, metaTagList)
    dictXMLWriter(excelPath, dictxmlPath, dictTagList)
    # print('\xa0')